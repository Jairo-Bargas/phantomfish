"""Genera la planilla Excel (mismo formato que el modelo) con datos reales."""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.constants import label_for
from app.money import ZERO, money
from app.models import Partner, Payment, Purchase, Sale, Settlement
from app.services.categories import label_map
from app.services.summary import build_summary

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F2937")
SUBTITLE_FONT = Font(italic=True, size=9, color="6B7280")
TOTAL_FONT = Font(bold=True, size=10)
OK_FILL = PatternFill("solid", fgColor="DCFCE7")
BAD_FILL = PatternFill("solid", fgColor="FEE2E2")
MONEY_FMT = '#,##0.00'
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _dec(value) -> float:
    return float(money(value))


def _flt(stmt, col, date_from, date_to):
    if date_from:
        stmt = stmt.where(col >= date_from)
    if date_to:
        stmt = stmt.where(col <= date_to)
    return stmt


def _header(ws, row: int, headers: list[str]) -> None:
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 28


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _sheet_instrucciones(wb: Workbook, generated: dt.datetime) -> None:
    ws = wb.active
    ws.title = "Instrucciones"
    ws["B2"] = "Phantom Fish — Gestión administrativa"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"Exportado el {generated:%d/%m/%Y %H:%M} desde la app."
    ws["B3"].font = SUBTITLE_FONT
    lines = [
        "",
        "Esta planilla es una FOTO de la base de datos al momento de exportar.",
        "Los montos están congelados: reflejan la cotización usada el día de cada pago.",
        "",
        "Hojas:",
        "  • Pagos_Aportes  — todo lo que sale de la empresa + cuánto puso cada socio.",
        "  • Costos_Compras — detalle de mercadería importada (se vincula por ID de pago).",
        "  • Ventas_Ingresos — cada venta (precio x cantidad).",
        "  • Movimientos_Socios — devoluciones / pagos de una parte al otro socio.",
        "  • Resumen_Socios — tablero: aportes, saldo neto entre socios y reparto.",
        "",
        "Control clave (Pagos_Aportes): la columna 'Total (ARS)' tiene que coincidir",
        "con 'Aportó Socio 1 + Aportó Socio 2'. Si no coincide, la fila queda marcada en rojo.",
    ]
    for i, line in enumerate(lines, start=5):
        ws.cell(row=i, column=2, value=line)
    _autofit(ws, [3, 90])


def _sheet_pagos(
    wb: Workbook, db: Session, partners: list[Partner], cat_labels: dict, date_from, date_to
) -> None:
    ws = wb.create_sheet("Pagos_Aportes")
    ws["A1"] = "PAGOS Y APORTES — Todo lo que sale de la empresa"
    ws["A1"].font = TITLE_FONT

    p_names = [p.name for p in partners]
    headers = [
        "ID", "Fecha", "Concepto", "Categoría", "Moneda", "Monto original",
        "Cotización", "Tipo", "Total (ARS)", "Total (USD)", "Estado",
    ]
    headers += [f"Corresponde {n} (ARS)" for n in p_names]
    headers += [f"Aportó {n} (ARS)" for n in p_names]
    headers += ["Suma aportes", "Control", "Notas"]
    _header(ws, 3, headers)

    payments = list(
        db.scalars(
            _flt(
                select(Payment).options(selectinload(Payment.contributions)),
                Payment.date, date_from, date_to,
            ).order_by(Payment.date, Payment.id)
        )
    )
    n = len(partners)
    row = 4
    for pay in payments:
        contrib = {c.partner_id: c.amount_ars for c in pay.contributions}
        corresponde = [
            _dec(money(pay.amount_ars) * p.pct_share / Decimal(100)) for p in partners
        ]
        aporto = [_dec(contrib.get(p.id, ZERO)) for p in partners]
        suma = money(sum((contrib.get(p.id, ZERO) for p in partners), ZERO))
        ok = abs(money(pay.amount_ars) - suma) <= Decimal("0.01")

        values = [
            f"ID{pay.id:04d}", pay.date, pay.concept,
            cat_labels.get(pay.category, pay.category), pay.currency_charged,
            _dec(pay.amount_original), float(pay.exchange_rate),
            label_for("rate_type", pay.exchange_rate_type),
            _dec(pay.amount_ars), _dec(pay.amount_usd),
            label_for("payment_status", pay.status),
            *corresponde, *aporto, _dec(suma),
            "OK" if ok else "REVISAR", pay.notes or "",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = BOX
            if isinstance(val, float):
                c.number_format = MONEY_FMT
        ctrl = ws.cell(row=row, column=11 + 2 * n + 2)
        ctrl.fill = OK_FILL if ok else BAD_FILL
        ctrl.font = TOTAL_FONT
        if not ok:
            for col in range(1, len(values) + 1):
                ws.cell(row=row, column=col).fill = BAD_FILL
        row += 1

    # fila de totales
    total_col = 9  # Total (ARS)
    usd_col = 10
    suma_col = 11 + 2 * n + 1
    tcell = ws.cell(row=row + 1, column=3, value="TOTALES")
    tcell.font = TOTAL_FONT
    for col in (total_col, usd_col, *range(12, 12 + 2 * n), suma_col):
        letter = get_column_letter(col)
        cell = ws.cell(row=row + 1, column=col, value=f"=SUM({letter}4:{letter}{row - 1})")
        cell.font = TOTAL_FONT
        cell.number_format = MONEY_FMT

    widths = [10, 12, 34, 15, 8, 14, 12, 10, 15, 15, 11] + [18] * (2 * n) + [15, 10, 30]
    _autofit(ws, widths)
    ws.freeze_panes = "C4"


def _sheet_compras(wb: Workbook, db: Session, date_from, date_to) -> None:
    ws = wb.create_sheet("Costos_Compras")
    ws["A1"] = "COSTOS / COMPRAS AL PROVEEDOR (detalle de mercadería)"
    ws["A1"].font = TITLE_FONT
    headers = [
        "ID", "Fecha", "Proveedor", "N° Factura", "ID Pago", "Producto / ítem",
        "Cantidad", "Precio unit. (USD)", "Total (USD)", "Estado envío", "Notas",
    ]
    _header(ws, 3, headers)

    purchases = list(
        db.scalars(
            _flt(
                select(Purchase).options(selectinload(Purchase.items)),
                Purchase.date, date_from, date_to,
            ).order_by(Purchase.date, Purchase.id)
        )
    )
    row = 4
    for pur in purchases:
        items = pur.items or [None]
        for it in items:
            values = [
                f"C{pur.id:04d}", pur.date, pur.supplier, pur.invoice_number or "",
                f"ID{pur.payment_id:04d}" if pur.payment_id else "",
                it.product_name if it else "(sin detalle)",
                float(it.quantity) if it else None,
                _dec(it.unit_price_usd) if it else None,
                _dec(it.total_usd) if it else None,
                label_for("shipment_status", pur.shipment_status),
                pur.notes or "",
            ]
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = BOX
                if col in (8, 9) and isinstance(val, float):
                    c.number_format = MONEY_FMT
            row += 1

    ws.cell(row=row + 1, column=8, value="TOTAL USD").font = TOTAL_FONT
    tc = ws.cell(row=row + 1, column=9, value=f"=SUM(I4:I{row - 1})")
    tc.font = TOTAL_FONT
    tc.number_format = MONEY_FMT
    _autofit(ws, [10, 12, 26, 14, 10, 34, 10, 16, 14, 16, 26])
    ws.freeze_panes = "C4"


def _sheet_ventas(wb: Workbook, db: Session, date_from, date_to) -> None:
    ws = wb.create_sheet("Ventas_Ingresos")
    ws["A1"] = "VENTAS / INGRESOS"
    ws["A1"].font = TITLE_FONT
    headers = [
        "ID", "Fecha", "Cliente / Canal", "Producto", "Cantidad",
        "Precio unit. (ARS)", "Total (ARS)", "Medio de pago", "Estado", "Notas",
    ]
    _header(ws, 3, headers)

    sales = list(
        db.scalars(
            _flt(
                select(Sale).options(selectinload(Sale.items)), Sale.date, date_from, date_to
            ).order_by(Sale.date, Sale.id)
        )
    )
    row = 4
    for sale in sales:
        canal = " / ".join(x for x in [sale.customer, label_for("sale_channel", sale.channel)] if x)
        items = sale.items or [None]
        for it in items:
            values = [
                f"V{sale.id:04d}", sale.date, canal,
                it.product_name if it else "(sin detalle)",
                float(it.quantity) if it else None,
                _dec(it.unit_price_ars) if it else None,
                _dec(it.total_ars) if it else None,
                label_for("payment_method", sale.payment_method),
                label_for("sale_status", sale.status),
                sale.notes or "",
            ]
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.border = BOX
                if col in (6, 7) and isinstance(val, float):
                    c.number_format = MONEY_FMT
            row += 1

    ws.cell(row=row + 1, column=6, value="TOTAL VENTAS").font = TOTAL_FONT
    tc = ws.cell(row=row + 1, column=7, value=f"=SUM(G4:G{row - 1})")
    tc.font = TOTAL_FONT
    tc.number_format = MONEY_FMT
    _autofit(ws, [10, 12, 28, 30, 10, 16, 14, 16, 12, 26])
    ws.freeze_panes = "C4"


def _sheet_resumen(wb: Workbook, db: Session, partners: list[Partner], date_from, date_to) -> None:
    ws = wb.create_sheet("Resumen_Socios")
    summary = build_summary(db, date_from=date_from, date_to=date_to)
    ws["B2"] = "RESUMEN DE SOCIOS Y REPARTO"
    ws["B2"].font = TITLE_FONT

    r = 4
    ws.cell(row=r, column=2, value="Porcentajes de reparto").font = TOTAL_FONT
    for p in partners:
        r += 1
        ws.cell(row=r, column=2, value=f"% {p.name}")
        pc = ws.cell(row=r, column=4, value=float(p.pct_share) / 100)
        pc.number_format = "0.00%"

    r += 3
    ws.cell(row=r, column=2, value="APORTES (ARS)").font = TOTAL_FONT
    r += 1
    ws.cell(row=r, column=2, value="")
    for i, ps in enumerate(summary.partners):
        ws.cell(row=r, column=3 + i, value=ps.name).font = TOTAL_FONT
    labels = [
        ("Le correspondía aportar", "should_contribute"),
        ("Aportó realmente", "did_contribute"),
        ("Diferencia (aportó − correspondía)", "balance"),
    ]
    for text_label, attr in labels:
        r += 1
        ws.cell(row=r, column=2, value=text_label)
        for i, ps in enumerate(summary.partners):
            c = ws.cell(row=r, column=3 + i, value=_dec(getattr(ps, attr)))
            c.number_format = MONEY_FMT

    r += 2
    ws.cell(row=r, column=2, value="Las devoluciones entre socios se registran aparte "
            "(hoja Movimientos_Socios) — no se descuentan acá.").font = SUBTITLE_FONT

    r += 3
    ws.cell(row=r, column=2, value="RESULTADO DEL PERÍODO").font = TOTAL_FONT
    rows = [
        ("Total ingresos (ventas)", _dec(summary.total_sales_ars)),
        ("Total pagos / costos", _dec(summary.total_payments_ars)),
        ("Ganancia / pérdida neta", _dec(summary.net_result_ars)),
    ]
    for text_label, val in rows:
        r += 1
        ws.cell(row=r, column=2, value=text_label)
        c = ws.cell(row=r, column=4, value=val)
        c.number_format = MONEY_FMT
    for ps in summary.partners:
        r += 1
        ws.cell(row=r, column=2, value=f"Corresponde a {ps.name} (según %)")
        c = ws.cell(row=r, column=4, value=_dec(ps.profit_share))
        c.number_format = MONEY_FMT

    r += 3
    ws.cell(row=r, column=2, value="CONTROL GLOBAL").font = TOTAL_FONT
    r += 1
    ws.cell(row=r, column=2, value="Suma de todos los pagos (ARS)")
    ws.cell(row=r, column=4, value=_dec(summary.total_payments_ars)).number_format = MONEY_FMT
    r += 1
    ws.cell(row=r, column=2, value="Suma de todos los aportes (ARS)")
    ws.cell(row=r, column=4, value=_dec(summary.total_contributions_ars)).number_format = MONEY_FMT
    r += 1
    ws.cell(row=r, column=2, value="Diferencia (debe ser 0)")
    diff = ws.cell(row=r, column=4, value=_dec(summary.control_difference))
    diff.number_format = MONEY_FMT
    diff.fill = OK_FILL if summary.control_ok else BAD_FILL
    diff.font = TOTAL_FONT

    _autofit(ws, [3, 40, 18, 18, 18])


def _sheet_movimientos(wb: Workbook, db: Session, date_from, date_to) -> None:
    ws = wb.create_sheet("Movimientos_Socios")
    ws["A1"] = "MOVIMIENTOS ENTRE SOCIOS (devoluciones / pagos de una parte al otro)"
    ws["A1"].font = TITLE_FONT
    _header(ws, 3, ["ID", "Fecha", "Pagó", "Recibió", "Monto (ARS)", "Medio", "Concepto", "Notas"])

    rows = list(
        db.scalars(
            _flt(select(Settlement), Settlement.date, date_from, date_to).order_by(
                Settlement.date, Settlement.id
            )
        )
    )
    row = 4
    for s in rows:
        values = [
            f"M{s.id:04d}", s.date,
            s.from_partner.name if s.from_partner else "",
            s.to_partner.name if s.to_partner else "",
            _dec(s.amount_ars), label_for("settlement_method", s.method),
            s.concept or "", s.notes or "",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = BOX
            if col == 5 and isinstance(val, float):
                c.number_format = MONEY_FMT
        row += 1
    ws.cell(row=row + 1, column=4, value="TOTAL").font = TOTAL_FONT
    tc = ws.cell(row=row + 1, column=5, value=f"=SUM(E4:E{row - 1})")
    tc.font = TOTAL_FONT
    tc.number_format = MONEY_FMT
    _autofit(ws, [10, 12, 16, 16, 16, 16, 30, 30])
    ws.freeze_panes = "C4"


def build_workbook(db: Session, *, date_from=None, date_to=None) -> bytes:
    generated = dt.datetime.now()
    partners = list(db.scalars(select(Partner).order_by(Partner.id)))
    cat_labels = label_map(db)
    wb = Workbook()
    _sheet_instrucciones(wb, generated)
    _sheet_pagos(wb, db, partners, cat_labels, date_from, date_to)
    _sheet_compras(wb, db, date_from, date_to)
    _sheet_ventas(wb, db, date_from, date_to)
    _sheet_movimientos(wb, db, date_from, date_to)
    _sheet_resumen(wb, db, partners, date_from, date_to)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def filename(date_from=None, date_to=None) -> str:
    if date_from or date_to:
        a = date_from.isoformat() if date_from else "inicio"
        b = date_to.isoformat() if date_to else "hoy"
        return f"phantomfish_{a}_a_{b}.xlsx"
    return f"phantomfish_gestion_{dt.date.today():%Y-%m-%d}.xlsx"
