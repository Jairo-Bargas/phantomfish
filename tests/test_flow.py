"""Prueba de extremo a extremo de los flujos principales."""

from __future__ import annotations

import io

from openpyxl import load_workbook


def test_login_required(client):
    resp = client.get("/", follow_redirects=False)
    assert resp.status_code == 303
    assert resp.headers["location"] == "/login"


def test_create_payment_auto_split_and_control(auth_client):
    resp = auth_client.post(
        "/pagos",
        data={
            "concept": "Pago proveedor China #45",
            "date": "2026-08-27",
            "category": "importacion",
            "status": "pagado",
            "currency_charged": "USD",
            "amount_original": "1000",
            "exchange_rate": "1300",
            "exchange_rate_type": "oficial",
            "split_mode": "auto",
        },
        follow_redirects=True,
    )
    assert resp.status_code == 200
    body = resp.text
    # total ARS = 1.300.000; split 35/65 -> 455.000 / 845.000
    assert "1.300.000" in body
    assert "455.000" in body
    assert "845.000" in body
    assert "coinciden con el total" in body


def test_create_payment_custom_split_mismatch_warns(auth_client):
    resp = auth_client.post(
        "/pagos",
        data={
            "concept": "Gasto operativo con error",
            "date": "2026-08-27",
            "category": "gasto_operativo",
            "status": "pagado",
            "currency_charged": "ARS",
            "amount_original": "100000",
            "exchange_rate": "1000",
            "exchange_rate_type": "oficial",
            "split_mode": "custom",
            "contribution_1": "10000",
            "contribution_2": "50000",
        },
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert "no coincide" in resp.text.lower() or "diferencia" in resp.text.lower()


def test_purchase_and_sale_totals(auth_client):
    auth_client.post(
        "/compras",
        data={
            "supplier": "Yiwu Fishing Co.",
            "date": "2026-08-20",
            "shipment_status": "en_transito",
            "item_name_0": "Señuelo cuchara 7cm",
            "item_qty_0": "500",
            "item_price_0": "1.8",
        },
        follow_redirects=True,
    )
    sale = auth_client.post(
        "/ventas",
        data={
            "date": "2026-08-25",
            "customer": "Pesca Total",
            "channel": "mayorista",
            "payment_method": "transferencia",
            "status": "cobrado",
            "item_name_0": "Señuelo cuchara 7cm",
            "item_qty_0": "50",
            "item_price_0": "3500",
        },
        follow_redirects=True,
    )
    assert "175.000" in sale.text  # 50 * 3500


def test_excel_export(auth_client):
    resp = auth_client.get("/export/excel")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == [
        "Instrucciones",
        "Pagos_Aportes",
        "Costos_Compras",
        "Ventas_Ingresos",
        "Movimientos_Socios",
        "Resumen_Socios",
    ]


def test_excel_export_by_month(auth_client):
    resp = auth_client.get("/export/excel?mes=2026-08")
    assert resp.status_code == 200
    assert "2026-08" in resp.headers["content-disposition"]


def test_data_browser(auth_client):
    resp = auth_client.get("/datos?tabla=payments")
    assert resp.status_code == 200
    assert "payments" in resp.text
    assert "password_hash" not in resp.text
